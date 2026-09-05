"""Admin chart lines: Excel round-trip + the user-side resolver.

The admin picks a segment, downloads a workbook pre-filled with every
instrument in it, types up to ``MAX_LEVELS`` price/colour pairs per row, and
uploads it back. Each price becomes a horizontal line on that instrument's
chart in the colour that was given.

Ownership and the user-side cascade mirror ``crypto_config_service`` exactly —
the same resolver rule already backs company banks and crypto configs, and a
second, subtly different one would be a bug waiting to happen.
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from beanie import PydanticObjectId
from bson import Decimal128
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.chart_level import ChartLevel, ChartLevelEntry
from app.models.instrument import Instrument
from app.models.user import User, UserRole
from app.utils.decimal_utils import to_decimal

MAX_LEVELS = 4

# Readable defaults so a sheet filled in without touching the colour columns
# still yields four DISTINGUISHABLE lines rather than four identical ones.
DEFAULT_COLORS = ["#E31E24", "#0EA5E9", "#16A34A", "#F59E0B"]

_HEX_RE = re.compile(r"^#(?:[0-9A-Fa-f]{3}|[0-9A-Fa-f]{6})$")

# Colour words an operator is likely to type instead of a hex code.
_NAMED: dict[str, str] = {
    "red": "#E31E24", "green": "#16A34A", "blue": "#2563EB", "yellow": "#EAB308",
    "orange": "#F59E0B", "purple": "#7C3AED", "violet": "#7C3AED", "pink": "#EC4899",
    "cyan": "#06B6D4", "sky": "#0EA5E9", "black": "#111111", "white": "#FFFFFF",
    "grey": "#6B7280", "gray": "#6B7280", "brown": "#92400E", "teal": "#14B8A6",
    "lime": "#84CC16", "magenta": "#D946EF", "gold": "#D4AF37", "silver": "#C0C0C0",
}

HEADERS = ["Token", "Symbol", "Segment"]
for _i in range(1, MAX_LEVELS + 1):
    HEADERS += [f"Price {_i}", f"Color {_i}", f"Label {_i}"]


# ── ownership ────────────────────────────────────────────────────────
def owner_filter_for_admin(admin: User) -> dict[str, PydanticObjectId | None]:
    """The (owner_admin_id, owner_broker_id) pair identifying THIS actor's
    rows. Super-admin writes the platform default (both null)."""
    if admin.role == UserRole.SUPER_ADMIN:
        return {"owner_admin_id": None, "owner_broker_id": None}
    if admin.role == UserRole.BROKER:
        return {"owner_admin_id": None, "owner_broker_id": admin.id}
    return {"owner_admin_id": admin.id, "owner_broker_id": None}


def normalize_color(raw: Any, fallback: str) -> str:
    """Accept '#RGB', '#RRGGBB', bare hex, or a colour name.

    Anything unrecognised falls back rather than passing through: an invalid
    CSS colour reaches TradingView as a black or invisible line with no error
    anywhere, which reads as "the feature is broken".
    """
    s = str(raw or "").strip()
    if not s:
        return fallback
    if _HEX_RE.match(s):
        return s.upper()
    if _HEX_RE.match("#" + s):
        return ("#" + s).upper()
    return _NAMED.get(s.lower(), fallback)


# ── template ─────────────────────────────────────────────────────────
async def build_template(admin: User, segment: str) -> bytes:
    """Workbook of every instrument in `segment`, pre-filled with whatever
    this admin already saved so editing is a round-trip, not a retype."""
    instruments = await Instrument.find(Instrument.segment == segment).to_list()
    instruments.sort(key=lambda i: (i.symbol or ""))

    f = owner_filter_for_admin(admin)
    existing = {
        r.token: r
        for r in await ChartLevel.find(
            ChartLevel.owner_admin_id == f["owner_admin_id"],
            ChartLevel.owner_broker_id == f["owner_broker_id"],
            ChartLevel.segment == segment,
        ).to_list()
    }

    wb = Workbook()
    ws = wb.active
    ws.title = (segment or "Levels")[:31]

    head_fill = PatternFill("solid", fgColor="0B1220")
    head_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = head_fill, head_font
        cell.alignment = Alignment(horizontal="center")

    for r, inst in enumerate(instruments, start=2):
        ws.cell(row=r, column=1, value=inst.token)
        ws.cell(row=r, column=2, value=inst.symbol)
        ws.cell(row=r, column=3, value=str(inst.segment))
        saved = existing.get(inst.token)
        for i in range(MAX_LEVELS):
            base = 4 + i * 3
            entry = saved.levels[i] if saved and i < len(saved.levels) else None
            if entry is not None:
                ws.cell(row=r, column=base, value=float(entry.price.to_decimal()))
                ws.cell(row=r, column=base + 1, value=entry.color)
                ws.cell(row=r, column=base + 2, value=entry.label or "")
            else:
                # Price stays blank (blank = no line) but the colour is seeded,
                # so typing only a price still gives four distinct colours.
                ws.cell(row=r, column=base + 1, value=DEFAULT_COLORS[i])

    for c in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"

    # A second sheet documenting the columns — whoever fills this in is not
    # the person who wrote the parser.
    guide = wb.create_sheet("How to fill")
    rows = [
        ("Column", "What to put"),
        ("Token / Symbol / Segment", "Leave as-is. Token identifies the instrument."),
        ("Price 1..4", "The price to draw a line at. Leave blank for no line."),
        ("Color 1..4", "Hex like #E31E24, or a name: red, green, blue, orange, purple, ..."),
        ("Label 1..4", "Optional text shown on the line."),
        ("", ""),
        ("Note", "Re-uploading REPLACES that instrument's lines with what the sheet says."),
        ("Note", "A row with every price blank CLEARS that instrument's lines."),
    ]
    for r, (a, b) in enumerate(rows, start=1):
        guide.cell(row=r, column=1, value=a).font = Font(bold=(r == 1))
        guide.cell(row=r, column=2, value=b).font = Font(bold=(r == 1))
    guide.column_dimensions["A"].width = 26
    guide.column_dimensions["B"].width = 78

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── import ───────────────────────────────────────────────────────────
async def import_workbook(admin: User, data: bytes) -> dict[str, Any]:
    """Parse an uploaded template and upsert this admin's rows.

    Collects problems instead of raising on the first bad cell — an operator
    with forty rows wants every error at once, not one per upload.
    """
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Not a readable .xlsx file: {exc}") from exc
    ws = wb.worksheets[0]

    f = owner_filter_for_admin(admin)
    known = {i.token: i for i in await Instrument.find().to_list()}

    updated = cleared = 0
    errors: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        token = str(row[0]).strip()
        inst = known.get(token)
        if inst is None:
            errors.append(f"{token}: not an instrument on this platform - skipped")
            continue

        entries: list[ChartLevelEntry] = []
        for i in range(MAX_LEVELS):
            base = 3 + i * 3  # values_only rows are 0-indexed
            price_raw = row[base] if len(row) > base else None
            if price_raw in (None, ""):
                continue
            try:
                price = to_decimal(price_raw)
            except Exception:  # noqa: BLE001
                errors.append(f"{inst.symbol}: 'Price {i + 1}' = {price_raw!r} is not a number")
                continue
            if price <= 0:
                errors.append(f"{inst.symbol}: 'Price {i + 1}' must be greater than 0")
                continue
            color_raw = row[base + 1] if len(row) > base + 1 else None
            label_raw = row[base + 2] if len(row) > base + 2 else None
            entries.append(
                ChartLevelEntry(
                    price=Decimal128(str(price)),
                    color=normalize_color(color_raw, DEFAULT_COLORS[i]),
                    label=(str(label_raw).strip() or None) if label_raw else None,
                )
            )

        doc = await ChartLevel.find_one(
            ChartLevel.owner_admin_id == f["owner_admin_id"],
            ChartLevel.owner_broker_id == f["owner_broker_id"],
            ChartLevel.token == token,
        )
        if not entries:
            # Every price blank = clear this instrument's lines.
            if doc is not None:
                await doc.delete()
                cleared += 1
            continue
        if doc is None:
            doc = ChartLevel(
                owner_admin_id=f["owner_admin_id"],
                owner_broker_id=f["owner_broker_id"],
                token=token,
                symbol=inst.symbol,
                segment=str(inst.segment),
            )
        doc.symbol = inst.symbol
        doc.segment = str(inst.segment)
        doc.levels = entries
        await doc.save()
        updated += 1

    return {"updated": updated, "cleared": cleared, "errors": errors}


# ── reads ────────────────────────────────────────────────────────────
def to_dict(doc: ChartLevel) -> dict[str, Any]:
    return {
        "token": doc.token,
        "symbol": doc.symbol,
        "segment": doc.segment,
        "levels": [
            {
                "price": float(e.price.to_decimal()),
                "color": e.color,
                "label": e.label,
            }
            for e in doc.levels
        ],
    }


async def list_for_admin(admin: User, segment: str | None = None) -> list[dict[str, Any]]:
    f = owner_filter_for_admin(admin)
    q = ChartLevel.find(
        ChartLevel.owner_admin_id == f["owner_admin_id"],
        ChartLevel.owner_broker_id == f["owner_broker_id"],
    )
    if segment:
        q = q.find(ChartLevel.segment == segment)
    return [to_dict(d) for d in await q.to_list()]


async def clear_for_admin(admin: User, token: str) -> bool:
    f = owner_filter_for_admin(admin)
    doc = await ChartLevel.find_one(
        ChartLevel.owner_admin_id == f["owner_admin_id"],
        ChartLevel.owner_broker_id == f["owner_broker_id"],
        ChartLevel.token == token,
    )
    if doc is None:
        return False
    await doc.delete()
    return True


async def resolve_for_user(user: User, token: str) -> list[dict[str, Any]]:
    """Lines the USER should see for `token`. Closest owner in the cascade
    wins — identical ordering to ``crypto_config_service.resolve_for_user``."""
    tried: list[dict[str, PydanticObjectId | None]] = []
    if user.assigned_broker_id is not None:
        tried.append({"owner_admin_id": None, "owner_broker_id": user.assigned_broker_id})
    for parent in reversed(list(user.broker_ancestry or [])):
        if parent == user.assigned_broker_id:
            continue
        tried.append({"owner_admin_id": None, "owner_broker_id": parent})
    if user.assigned_admin_id is not None:
        tried.append({"owner_admin_id": user.assigned_admin_id, "owner_broker_id": None})
    else:
        # Direct super-admin user only — never leak the platform default to a
        # sub-admin's users (same rule as company banks).
        tried.append({"owner_admin_id": None, "owner_broker_id": None})

    for f in tried:
        doc = await ChartLevel.find_one(
            ChartLevel.owner_admin_id == f["owner_admin_id"],
            ChartLevel.owner_broker_id == f["owner_broker_id"],
            ChartLevel.token == token,
        )
        if doc is not None and doc.levels:
            return to_dict(doc)["levels"]
    return []
